import openpyxl

# Create a new workbook
wb = openpyxl.Workbook()

# Remove the default sheet
wb.remove(wb.active)

# M Summary sheet
ws = wb.create_sheet("M Summary")
data_summary = [
    ["", "", "", "", "", "", ""],
    ["Month - Jan 26", "", "", "", "", "", ""],
    ["Prev. Bal to be reconciled", 928947, "", "", "", "", ""],
    ["", 178470, "", "", "", "", ""],
    ["", 0, "Total Amount Received", "='Received Amount'!E27", "", "", "Reconsiled = Bills received"],
    ["", 0, "Total Amount Issued", "='Summary Payment'!C14", "", "", ""],
    ["", 928947, "Total Amount Reconsiled", "='Summary Payment'!D14", "", "", ""],
    ["D = B - C", "Bal to be reconsiled", "", "=D5-D6", "", "", ""],
    ["", 178470, "", "", "", "", ""],
    ["E = A - B", "Cash in Hand", "", "=D4-D5", "", "", ""],
]
for row in data_summary:
    ws.append(row)

# Received Amount sheet
ws = wb.create_sheet("Received Amount")
data_received = [
    ["RECEIVED TO PETTY CASH - DETAILS", "", "", "", "", "", ""],
    ["Ser", "Date", "Received From", "Cheuqe #", "Amount", "Head", "Remarks"],
    [1, 46023, "BBF", "", 178470, "", "From Dec 2025 Cash in Hand"],
]
for _ in range(23):  # Blank rows to match structure
    data_received.append(["", "", "", "", "", "", ""])
data_received.append(["Total", "", "", "", 178470, "", ""])
for row in data_received:
    ws.append(row)

# Summary Payment sheet
ws = wb.create_sheet("Summary Payment")
ws.append(["SUMMERY OF DISTRIBUTED PETTY CASH ", "", "", "", "", ""])
ws.append(["Ser", "Name", "Prev Bal", "Issue Amount", "Reconsiled Amount", "Bal Amount"])
recip_data = [
    ["1", "DM Awais", 0, 0, 0, 0],
    ["2", "AM Awais (OFC Sup)", -33683, 0, 0, -33683, "Reconsiled = Bills received"],
    ["3", "AM Coord", 0, 0, 0, 0],
    ["4", "AM Irfan", 0, 0, 0, 0],
    ["5", "AM Ahmed", 100000, 0, 0, 100000],
    ["6", "JM Mubashir", 110000, 0, 0, 110000],
    ["7", "JM Adil", 150000, 0, 0, 150000],
    ["8", "JM Waheed", 25980, 0, 0, 25980],
    ["9", "JM Haseeb", 0, 0, 0, 0],
    ["10", "Adm Expenses", 426650, 0, 0, 426650],
    ["11", "AM Essa Khan ", 150000, 0, 0, 150000],
]
for row in recip_data:
    ws.append(row)
ws.append(["Total Bal Amount", "", 928947, 0, 0, 928947])

# Individual recipient sheets
individual_data = [
    ("1. DM Awais", 0),
    ("2. AM Awais (OFC)", -33683),
    ("3. AM Coord", 0),
    ("4. AM Irfan", 0),
    ("5. AM Ahmed", 100000),
    ("6.JM Mubashir", 110000),
    ("7. JM Adil", 150000),
    ("8. JM Waheed", 25980),
    ("9. JM Haseeb", 0),
    ("10.Adm Expenses", 426650),
    ("11. AM Essa Khan ", 150000),
]

for sheet_name, prev_bal in individual_data:
    ws = wb.create_sheet(sheet_name)
    ws.append(["PREVIOUS BALANCE"] + [""] * 9)
    ws.append(["Prev Bal to Reconcile", "", prev_bal] + [""] * 7)
    ws.append(["ISSUED AMOUNT"] + [""] * 9)
    ws.append(["Ser", "Dated", "Amount", "Remarks"] + [""] * 6)
    for _ in range(8):  # Blank issued rows
        ws.append([""] * 10)
    total_label = "Total Issued" + (" (Adm Expenses)" if "Adm" in sheet_name else (" (AM Essa Khan)" if "Essa" in sheet_name else ""))
    ws.append([total_label, "", 0] + [""] * 7)
    ws.append([""] * 10)
    ws.append(["RECONCILLED AMOUNT"] + [""] * 5)
    ws.append(["", "", "", "", "Ser", "Dated", "Nomenclature", "Expense Head", "Details of Expance", "Amount"])
    for _ in range(10):  # Blank reconciled rows
        ws.append([""] * 10)
    ws.append(["", "", "", "", "Total Reconcilled", "", "", "", "", 0])

# Save the workbook
wb.save("Petty_Cash_Jan_26.xlsx")
print("Workbook generated successfully: Petty_Cash_Jan_26.xlsx")